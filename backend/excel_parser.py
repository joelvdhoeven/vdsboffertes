"""
Excel prijzenboek parser
Extracts all werkzaamheden from the Prijzenboek sheet
"""
import openpyxl
from typing import List, Dict, Any


def parse_prijzenboek(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse the Prijzenboek sheet from Excel file
    Returns list of werkzaamheden with code, description, unit, and prices
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    if "Prijzenboek" not in wb.sheetnames:
        raise ValueError("Prijzenboek sheet not found in Excel file")

    sheet = wb["Prijzenboek"]

    prijzenboek = []

    # Headers are in row 1
    # Row 1: CODERING DATABASE | OMSCHRIJVING | EENHEID | MATERIAAL EXCL BTW | UREN EXCL BTW | TOTAAL EXCL BTW | TOTAAL INCL BTW | BTW LAAG MAT

    # Start from row 2 (skip header)
    for row_num in range(2, sheet.max_row + 1):
        # Get values
        code = sheet.cell(row=row_num, column=1).value  # Column A: CODERING
        omschrijving = sheet.cell(row=row_num, column=2).value  # Column B: OMSCHRIJVING
        eenheid = sheet.cell(row=row_num, column=3).value  # Column C: EENHEID
        materiaal_excl = sheet.cell(row=row_num, column=4).value  # Column D: MATERIAAL EXCL BTW
        uren_excl = sheet.cell(row=row_num, column=5).value  # Column E: UREN EXCL BTW
        totaal_excl = sheet.cell(row=row_num, column=6).value  # Column F: TOTAAL EXCL BTW
        totaal_incl = sheet.cell(row=row_num, column=7).value  # Column G: TOTAAL INCL BTW
        btw_laag = sheet.cell(row=row_num, column=8).value  # Column H: BTW LAAG MAT

        # Skip empty rows or category headers (rows without description)
        if not omschrijving or omschrijving.strip() == "":
            continue

        # Skip rows that are just category headers (no eenheid)
        # Category headers have codes but no prices/units
        if not eenheid:
            continue

        # Normalize eenheid
        if eenheid:
            eenheid = str(eenheid).strip().lower()

        item = {
            "code": str(code) if code else "",
            "omschrijving": str(omschrijving).strip(),
            "eenheid": eenheid,
            "materiaal_excl": float(materiaal_excl) if materiaal_excl else 0.0,
            "uren_excl": float(uren_excl) if uren_excl else 0.0,
            "totaal_excl": float(totaal_excl) if totaal_excl else 0.0,
            "totaal_incl": float(totaal_incl) if totaal_incl else 0.0,
            "btw_laag": float(btw_laag) if btw_laag else 0.0,
            "row_num": row_num  # Store original row number for reference
        }

        prijzenboek.append(item)

    wb.close()

    return prijzenboek


def find_prijzenboek_by_code(prijzenboek: List[Dict[str, Any]], code: str) -> Dict[str, Any]:
    """Find a prijzenboek item by its code"""
    for item in prijzenboek:
        if item["code"] == code:
            return item
    return None


def search_prijzenboek(prijzenboek: List[Dict[str, Any]], query: str, limit: int = 10) -> List[Dict[str, Any]]:
    """
    Simple search in prijzenboek by description
    Returns top matches
    """
    query_lower = query.lower()
    matches = []

    for item in prijzenboek:
        if query_lower in item["omschrijving"].lower():
            matches.append(item)

            if len(matches) >= limit:
                break

    return matches


if __name__ == "__main__":
    # Test with example file
    prijzenboek = parse_prijzenboek("/home/user/vdsboffertes/Opnamelijst_-_Woonforte_2025_22-04__Nieuw_.xlsm")

    print("=" * 80)
    print("PRIJZENBOEK PARSED")
    print("=" * 80)
    print(f"\nTotal items: {len(prijzenboek)}")

    print("\n\nFirst 10 items:")
    print("-" * 80)
    for i, item in enumerate(prijzenboek[:10], 1):
        print(f"\n{i}. Code: {item['code']}")
        print(f"   Omschrijving: {item['omschrijving']}")
        print(f"   Eenheid: {item['eenheid']}")
        print(f"   Prijs excl: €{item['totaal_excl']:.2f}")
        print(f"   Prijs incl: €{item['totaal_incl']:.2f}")

    # Test search
    print("\n\n" + "=" * 80)
    print("SEARCH TEST: 'behang'")
    print("=" * 80)

    results = search_prijzenboek(prijzenboek, "behang", limit=5)
    for i, item in enumerate(results, 1):
        print(f"\n{i}. {item['omschrijving']} ({item['eenheid']}) - €{item['totaal_excl']:.2f}")

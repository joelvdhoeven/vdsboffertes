"""
Script om vereenvoudigde prijzenboek CSV te importeren in het Excel prijzenboek
CSV formaat (puntkomma-gescheiden):
- CODERING DATABASE
- OMSCHRIJVING VAKMAN MUTATIE
- EENHEID
- Materiaal per stuk EXCL BTW
- Uren per stuk EXCL BTW
- Prijs per stuk EXCL BTW
- OMSCHRIJVING OFFERTE MUTATIE

Wordt toegevoegd aan Excel met alle ruimte kolommen (C-O) op 0
"""
import openpyxl
import csv
import sys
from typing import Dict, List


def parse_csv(csv_path: str) -> List[Dict]:
    """Parse CSV file met prijzenboek data"""
    items = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            # Clean en parse bedragen (verwijder € en komma's)
            def clean_amount(val):
                if not val:
                    return 0.0
                # Verwijder € symbool en spaties
                val = str(val).replace('€', '').replace(' ', '').strip()
                # Vervang komma door punt voor decimalen
                val = val.replace(',', '.')
                try:
                    return float(val)
                except ValueError:
                    return 0.0

            item = {
                'code': row['CODERING DATABASE'].strip(),
                'omschrijving': row['OMSCHRIJVING VAKMAN MUTATIE'].strip(),
                'eenheid': row['EENHEID'].strip(),
                'materiaal': clean_amount(row['Materiaal per stuk EXCL BTW']),
                'uren': clean_amount(row['Uren per stuk EXCL BTW']),
                'prijs_per_stuk': clean_amount(row['Prijs per stuk EXCL BTW']),
                'totaal_excl': 0.0,  # Not in simplified CSV
                'totaal_incl': 0.0,  # Not in simplified CSV
                'omschrijving_offerte': row['OMSCHRIJVING OFFERTE MUTATIE'].strip(),
            }
            items.append(item)
    return items


def find_row_by_code(sheet, code: str) -> int:
    """Vind rij in Excel op basis van code. Return 0 als niet gevonden"""
    for row_num in range(2, sheet.max_row + 1):
        cell_code = sheet.cell(row=row_num, column=1).value
        if cell_code and str(cell_code).strip() == code:
            return row_num
    return 0


def add_or_update_row(sheet, item: Dict, row_num: int = None):
    """
    Voeg nieuwe rij toe of update bestaande rij
    Als row_num is None, voeg toe aan het einde
    """
    if row_num is None:
        row_num = sheet.max_row + 1
        action = "ADDED"
    else:
        action = "UPDATED"

    # Basis informatie
    sheet.cell(row=row_num, column=1).value = item['code']  # A
    sheet.cell(row=row_num, column=2).value = item['omschrijving']  # B

    # Ruimtes (C-O) - allemaal op 0
    for col in range(3, 16):  # C=3 tot O=15
        sheet.cell(row=row_num, column=col).value = 0.0

    # Kolom P leeglaten (16)

    # TOTAAL (Q) - bereken als som van alle ruimtes * prijs
    # Voor nu op 0 omdat alle ruimtes 0 zijn
    sheet.cell(row=row_num, column=17).value = 0.0  # Q

    # Prijzen
    sheet.cell(row=row_num, column=18).value = item['eenheid']  # R
    sheet.cell(row=row_num, column=19).value = item['materiaal']  # S
    sheet.cell(row=row_num, column=20).value = item['uren']  # T
    sheet.cell(row=row_num, column=21).value = item['prijs_per_stuk']  # U

    # Kolom V leeglaten (22)

    sheet.cell(row=row_num, column=23).value = item['totaal_excl']  # W
    sheet.cell(row=row_num, column=24).value = item['totaal_incl']  # X
    sheet.cell(row=row_num, column=25).value = item['omschrijving_offerte']  # Y

    return action


def import_csv_to_excel(csv_path: str, excel_path: str, backup: bool = True):
    """
    Importeer CSV data naar Excel prijzenboek

    Args:
        csv_path: Pad naar CSV bestand
        excel_path: Pad naar Excel prijzenboek
        backup: Maak backup van Excel bestand voor wijzigen
    """
    # Parse CSV
    print(f"Parsing CSV: {csv_path}")
    items = parse_csv(csv_path)
    print(f"Found {len(items)} items in CSV")

    # Maak backup
    if backup:
        import shutil
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = excel_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
        shutil.copy2(excel_path, backup_path)
        print(f"Backup created: {backup_path}")

    # Open Excel
    print(f"Opening Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Process items
    added = 0
    updated = 0

    for item in items:
        code = item['code']
        existing_row = find_row_by_code(sheet, code)

        if existing_row:
            action = add_or_update_row(sheet, item, existing_row)
            updated += 1
            print(f"  {action}: {code} - {item['omschrijving'][:50]}")
        else:
            action = add_or_update_row(sheet, item, None)
            added += 1
            print(f"  {action}: {code} - {item['omschrijving'][:50]}")

    # Save Excel
    wb.save(excel_path)
    wb.close()

    print(f"\nDone!")
    print(f"  Added: {added}")
    print(f"  Updated: {updated}")
    print(f"  Total: {len(items)}")


if __name__ == "__main__":
    # Default paths
    csv_path = "/home/user/vdsboffertes/backend/prijzenboek_import_template.csv"
    excel_path = "/home/user/vdsboffertes/backend/Juiste opnamelijst.xlsx"

    # Allow command line arguments
    if len(sys.argv) > 1:
        csv_path = sys.argv[1]
    if len(sys.argv) > 2:
        excel_path = sys.argv[2]

    print("=" * 80)
    print("PRIJZENBOEK CSV IMPORT")
    print("=" * 80)

    import_csv_to_excel(csv_path, excel_path, backup=True)

    print("\n" + "=" * 80)

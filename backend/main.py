"""
FastAPI backend for Offerte Generator MVP
"""
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import os
import sys
import shutil
from pathlib import Path
import uuid

# Add backend directory to path for imports
sys.path.insert(0, os.path.dirname(__file__))

# Import backend modules
try:
    # Try relative imports first (when running as package)
    from .document_parser import parse_docx_opname
    from .excel_parser import parse_prijzenboek
    from .matcher import match_werkzaamheden
    from .excel_generator import generate_filled_excel
except ImportError:
    # Fall back to absolute imports (when running directly)
    from document_parser import parse_docx_opname
    from excel_parser import parse_prijzenboek
    from matcher import match_werkzaamheden
    from excel_generator import generate_filled_excel

app = FastAPI(title="Offerte Generator API", version="1.0.0")

# CORS middleware for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, replace with specific origin
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Storage paths
UPLOAD_DIR = Path("/tmp/uploads") if os.getenv("RAILWAY_ENVIRONMENT") else Path("../uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# In-memory storage for sessions (in production, use Redis or DB)
sessions = {}


class MatchReview(BaseModel):
    """Model for match review data"""
    werkzaamheid_id: str
    prijzenboek_code: str
    accepted: bool


class GenerateRequest(BaseModel):
    """Model for generate request"""
    session_id: str
    matches: List[MatchReview]


@app.get("/")
async def root():
    """Health check endpoint"""
    return {"status": "ok", "message": "Offerte Generator API is running"}


@app.post("/api/upload/notes")
async def upload_notes(file: UploadFile = File(...)):
    """Upload Samsung Notes document (docx/txt)"""
    try:
        # Validate file type
        if not file.filename.endswith(('.docx', '.txt', '.pdf')):
            raise HTTPException(status_code=400, detail="Only .docx, .txt, and .pdf files are supported")

        # Create session
        session_id = str(uuid.uuid4())
        session_dir = UPLOAD_DIR / session_id
        session_dir.mkdir(exist_ok=True)

        # Save file
        notes_path = session_dir / f"notes_{file.filename}"
        with open(notes_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Store session
        sessions[session_id] = {
            "notes_path": str(notes_path),
            "prijzenboek_path": None,
            "parsed_opname": None,
            "prijzenboek_data": None,
            "matches": None
        }

        return {
            "session_id": session_id,
            "filename": file.filename,
            "message": "Notes uploaded successfully"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/upload/prijzenboek")
async def upload_prijzenboek(session_id: str, file: UploadFile = File(...)):
    """Upload Excel prijzenboek"""
    try:
        # Validate session
        if session_id not in sessions:
            raise HTTPException(status_code=404, detail="Session not found")

        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files are supported")

        # Save file
        session_dir = UPLOAD_DIR / session_id
        prijzenboek_path = session_dir / f"prijzenboek_{file.filename}"
        with open(prijzenboek_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Update session
        sessions[session_id]["prijzenboek_path"] = str(prijzenboek_path)

        return {
            "session_id": session_id,
            "filename": file.filename,
            "message": "Prijzenboek uploaded successfully"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/process/parse")
async def process_parse(session_id: str):
    """Parse uploaded documents"""
    try:
        # Validate session
        if session_id not in sessions:
            raise HTTPException(status_code=404, detail="Session not found")

        session = sessions[session_id]

        if not session["notes_path"]:
            raise HTTPException(status_code=400, detail="Notes not uploaded")

        # Parse opname
        parsed_opname = parse_docx_opname(session["notes_path"])
        session["parsed_opname"] = parsed_opname

        # Use default prijzenboek from admin if not uploaded for this session
        if not session["prijzenboek_path"]:
            default_prijzenboek_path = Path(__file__).parent / "Juiste opnamelijst.xlsx"
            if not default_prijzenboek_path.exists():
                raise HTTPException(status_code=404, detail="Default prijzenboek not found. Please upload via admin panel.")
            session["prijzenboek_path"] = str(default_prijzenboek_path)

        # Parse prijzenboek using new parser
        try:
            from .excel_parser_new import parse_prijzenboek_new
        except ImportError:
            from excel_parser_new import parse_prijzenboek_new

        prijzenboek_data = parse_prijzenboek_new(session["prijzenboek_path"])
        session["prijzenboek_data"] = prijzenboek_data

        # Count werkzaamheden
        total_werkzaamheden = sum(
            len(ruimte["werkzaamheden"])
            for ruimte in parsed_opname["ruimtes"]
        )

        return {
            "session_id": session_id,
            "parsed": True,
            "ruimtes": len(parsed_opname["ruimtes"]),
            "werkzaamheden": total_werkzaamheden,
            "prijzenboek_items": len(prijzenboek_data)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/process/match")
async def process_match(session_id: str):
    """Match werkzaamheden with prijzenboek"""
    try:
        # Validate session
        if session_id not in sessions:
            raise HTTPException(status_code=404, detail="Session not found")

        session = sessions[session_id]

        if not session["parsed_opname"] or not session["prijzenboek_data"]:
            raise HTTPException(status_code=400, detail="Documents not parsed yet")

        # Perform matching
        matches = match_werkzaamheden(
            session["parsed_opname"],
            session["prijzenboek_data"]
        )

        session["matches"] = matches

        # Calculate statistics
        high_confidence = sum(1 for m in matches if m["confidence"] >= 0.9)
        medium_confidence = sum(1 for m in matches if 0.7 <= m["confidence"] < 0.9)
        low_confidence = sum(1 for m in matches if m["confidence"] < 0.7)

        return {
            "session_id": session_id,
            "total_matches": len(matches),
            "high_confidence": high_confidence,
            "medium_confidence": medium_confidence,
            "low_confidence": low_confidence,
            "matches": matches
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/matches/update")
async def update_match(session_id: str, match_id: str, prijzenboek_code: str):
    """Update a specific match with a different prijzenboek item"""
    try:
        # Validate session
        if session_id not in sessions:
            raise HTTPException(status_code=404, detail="Session not found")

        session = sessions[session_id]

        if not session["matches"]:
            raise HTTPException(status_code=400, detail="No matches found")

        # Find and update the match
        for match in session["matches"]:
            if match["id"] == match_id:
                # Find the new prijzenboek item
                new_item = None
                for item in session["prijzenboek_data"]:
                    if item["code"] == prijzenboek_code:
                        new_item = item
                        break

                if not new_item:
                    raise HTTPException(status_code=404, detail="Prijzenboek item not found")

                # Update match
                match["prijzenboek_match"] = new_item
                match["confidence"] = 1.0  # Manual selection = 100% confidence
                match["match_type"] = "manual"

                return {"success": True, "match": match}

        raise HTTPException(status_code=404, detail="Match not found")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generate/excel")
async def generate_excel(request: GenerateRequest):
    """Generate filled Excel file"""
    try:
        # Validate session
        if request.session_id not in sessions:
            raise HTTPException(status_code=404, detail="Session not found")

        session = sessions[request.session_id]

        if not session["matches"]:
            raise HTTPException(status_code=400, detail="No matches found")

        # Generate Excel file
        output_path = generate_filled_excel(
            template_path=session["prijzenboek_path"],
            matches=session["matches"],
            session_dir=UPLOAD_DIR / request.session_id
        )

        session["output_excel"] = output_path

        return {
            "success": True,
            "file_path": output_path,
            "download_url": f"/api/download/excel/{request.session_id}"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/download/excel/{session_id}")
async def download_excel(session_id: str):
    """Download generated Excel file"""
    try:
        # Validate session
        if session_id not in sessions:
            raise HTTPException(status_code=404, detail="Session not found")

        session = sessions[session_id]

        if not session.get("output_excel"):
            raise HTTPException(status_code=404, detail="Excel file not generated yet")

        file_path = session["output_excel"]
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")

        return FileResponse(
            path=file_path,
            filename="Offerte_Generated.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/session/{session_id}/status")
async def get_session_status(session_id: str):
    """Get session status"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = sessions[session_id]

    return {
        "session_id": session_id,
        "notes_uploaded": session["notes_path"] is not None,
        "prijzenboek_uploaded": session["prijzenboek_path"] is not None,
        "parsed": session["parsed_opname"] is not None,
        "matched": session["matches"] is not None,
        "generated": session.get("output_excel") is not None
    }


# Admin endpoints
@app.get("/api/admin/prijzenboek")
async def get_prijzenboek_admin():
    """Get prijzenboek data for admin panel"""
    try:
        # Load the default prijzenboek file
        default_prijzenboek_path = Path(__file__).parent / "Juiste opnamelijst.xlsx"

        if not default_prijzenboek_path.exists():
            raise HTTPException(status_code=404, detail="Prijzenboek file not found")

        # Use new parser for the new format
        try:
            from .excel_parser_new import parse_prijzenboek_new
        except ImportError:
            from excel_parser_new import parse_prijzenboek_new

        prijzenboek_items = parse_prijzenboek_new(str(default_prijzenboek_path))

        return {
            "items": prijzenboek_items,
            "total": len(prijzenboek_items)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/prijzenboek")
async def save_prijzenboek_admin(data: Dict[str, Any]):
    """Save updated prijzenboek data"""
    try:
        import openpyxl

        # Load the workbook
        prijzenboek_path = Path(__file__).parent / "Juiste opnamelijst.xlsx"

        if not prijzenboek_path.exists():
            raise HTTPException(status_code=404, detail="Prijzenboek file not found")

        wb = openpyxl.load_workbook(str(prijzenboek_path))
        sheet = wb.active

        # Update existing rows and add new ones
        items = data.get("items", [])

        # Clear existing data (keep headers)
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None

        # Write updated data
        for idx, item in enumerate(items, start=2):
            sheet.cell(row=idx, column=1).value = item.get("code", "")
            sheet.cell(row=idx, column=2).value = item.get("omschrijving", "")
            sheet.cell(row=idx, column=18).value = item.get("eenheid", "")  # Column R
            sheet.cell(row=idx, column=19).value = item.get("materiaal", 0)  # Column S
            sheet.cell(row=idx, column=20).value = item.get("uren", 0)  # Column T

        # Save workbook
        wb.save(str(prijzenboek_path))
        wb.close()

        return {
            "success": True,
            "message": "Prijzenboek successfully updated",
            "items_saved": len(items)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/prijzenboek/upload")
async def upload_prijzenboek_admin(file: UploadFile = File(...)):
    """Upload and replace prijzenboek Excel file"""
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files are supported")

        # Save uploaded file
        prijzenboek_path = Path(__file__).parent / "Juiste opnamelijst.xlsx"

        # Backup old file
        backup_path = Path(__file__).parent / "Juiste opnamelijst_backup.xlsx"
        if prijzenboek_path.exists():
            shutil.copy(str(prijzenboek_path), str(backup_path))

        # Save new file
        with open(prijzenboek_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Parse the new file to verify it's valid
        try:
            from .excel_parser_new import parse_prijzenboek_new
        except ImportError:
            from excel_parser_new import parse_prijzenboek_new

        prijzenboek_items = parse_prijzenboek_new(str(prijzenboek_path))

        return {
            "success": True,
            "message": "Prijzenboek uploaded successfully",
            "items_loaded": len(prijzenboek_items),
            "filename": file.filename
        }

    except Exception as e:
        # Restore backup if upload failed
        if backup_path.exists():
            shutil.copy(str(backup_path), str(prijzenboek_path))
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

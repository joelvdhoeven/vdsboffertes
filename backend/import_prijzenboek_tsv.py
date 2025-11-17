"""
Script om TAB-gescheiden prijzenboek data te importeren in het Excel prijzenboek
Ondersteunt € notatie zoals "€ 6,285.20" of "-€ 249.78"

TSV formaat (TAB-gescheiden):
- CODERING DATABASE
- OMSCHRIJVING VAKMAN MUTATIE
- EENHEID
- Matriaal per stuk EXCL BTW (bijv. € 6,285.20)
- Uren per stuk EXCL BTW
- Prijs per stuk EXCL BTW
- TOTAAL EXCL BTW
- TOTAAL INCL BTW
- OMSCHRIJVING OFFERTE MUTATIE
"""
import openpyxl
import csv
import sys
import re
from typing import Dict, List


def clean_euro_amount(val):
    """
    Parse bedragen met € notatie
    Ondersteunt: "€ 6,285.20", "-€ 249.78", "€ 0.00", etc.
    """
    if not val:
        return 0.0

    val = str(val).strip()

    # Check for negative
    is_negative = '-' in val

    # Remove € symbol, spaces, and minus sign
    val = val.replace('€', '').replace(' ', '').replace('-', '').strip()

    # Handle European number format: 6,285.20 -> 6285.20
    # Remove thousands separator (comma in this case)
    val = val.replace(',', '')

    try:
        result = float(val)
        return -result if is_negative else result
    except ValueError:
        return 0.0


def parse_tsv(tsv_path: str) -> List[Dict]:
    """Parse TSV file met prijzenboek data"""
    items = []

    with open(tsv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter='\t')

        for row in reader:
            # Get columns - handle different possible column names
            code_col = 'CODERING DATABASE'
            omschrijving_col = 'OMSCHRIJVING VAKMAN MUTATIE'
            eenheid_col = 'EENHEID'

            # Try to find the right column names
            materiaal_col = None
            uren_col = None
            prijs_col = None
            totaal_excl_col = None
            totaal_incl_col = None
            offerte_col = None

            for key in row.keys():
                if 'Matriaal' in key or 'Materiaal' in key:
                    materiaal_col = key
                elif 'Uren per stuk' in key:
                    uren_col = key
                elif 'Prijs per stuk' in key:
                    prijs_col = key
                elif 'TOTAAL' in key and 'EXCL' in key:
                    totaal_excl_col = key
                elif 'TOTAAL' in key and 'INCL' in key:
                    totaal_incl_col = key
                elif 'OMSCHRIJVING OFFERTE' in key:
                    offerte_col = key

            if not materiaal_col:
                raise ValueError(f"Could not find materiaal column in: {list(row.keys())}")

            item = {
                'code': row[code_col].strip(),
                'omschrijving': row[omschrijving_col].strip(),
                'eenheid': row[eenheid_col].strip(),
                'materiaal': clean_euro_amount(row[materiaal_col]),
                'uren': clean_euro_amount(row[uren_col]) if uren_col else 0.0,
                'prijs_per_stuk': clean_euro_amount(row[prijs_col]) if prijs_col else 0.0,
                'totaal_excl': clean_euro_amount(row[totaal_excl_col]) if totaal_excl_col else 0.0,
                'totaal_incl': clean_euro_amount(row[totaal_incl_col]) if totaal_incl_col else 0.0,
                'omschrijving_offerte': row[offerte_col].strip() if offerte_col else row[omschrijving_col].strip(),
            }
            items.append(item)

    return items


def find_row_by_code(sheet, code: str) -> int:
    """Vind rij in Excel op basis van code. Return 0 als niet gevonden"""
    for row_num in range(2, sheet.max_row + 1):
        cell_code = sheet.cell(row=row_num, column=1).value
        if cell_code and str(cell_code).strip() == code:
            return row_num
    return 0


def add_or_update_row(sheet, item: Dict, row_num: int = None):
    """
    Voeg nieuwe rij toe of update bestaande rij
    Als row_num is None, voeg toe aan het einde
    """
    if row_num is None:
        row_num = sheet.max_row + 1
        action = "ADDED"
    else:
        action = "UPDATED"

    # Basis informatie
    sheet.cell(row=row_num, column=1).value = item['code']  # A
    sheet.cell(row=row_num, column=2).value = item['omschrijving']  # B

    # Ruimtes (C-O) - allemaal op 0
    for col in range(3, 16):  # C=3 tot O=15
        sheet.cell(row=row_num, column=col).value = 0.0

    # Kolom P leeglaten (16)

    # TOTAAL (Q) - bereken als som van alle ruimtes * prijs
    sheet.cell(row=row_num, column=17).value = 0.0  # Q

    # Prijzen
    sheet.cell(row=row_num, column=18).value = item['eenheid']  # R
    sheet.cell(row=row_num, column=19).value = item['materiaal']  # S
    sheet.cell(row=row_num, column=20).value = item['uren']  # T
    sheet.cell(row=row_num, column=21).value = item['prijs_per_stuk']  # U

    # Kolom V leeglaten (22)

    sheet.cell(row=row_num, column=23).value = item['totaal_excl']  # W
    sheet.cell(row=row_num, column=24).value = item['totaal_incl']  # X
    sheet.cell(row=row_num, column=25).value = item['omschrijving_offerte']  # Y

    return action


def import_tsv_to_excel(tsv_path: str, excel_path: str, backup: bool = True):
    """
    Importeer TSV data naar Excel prijzenboek

    Args:
        tsv_path: Pad naar TSV bestand
        excel_path: Pad naar Excel prijzenboek
        backup: Maak backup van Excel bestand voor wijzigen
    """
    # Parse TSV
    print(f"Parsing TSV: {tsv_path}")
    items = parse_tsv(tsv_path)
    print(f"Found {len(items)} items in TSV")

    # Maak backup
    if backup:
        import shutil
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = excel_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
        shutil.copy2(excel_path, backup_path)
        print(f"Backup created: {backup_path}")

    # Open Excel
    print(f"Opening Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Process items
    added = 0
    updated = 0

    for item in items:
        code = item['code']
        existing_row = find_row_by_code(sheet, code)

        if existing_row:
            action = add_or_update_row(sheet, item, existing_row)
            updated += 1
            print(f"  {action}: {code} - {item['omschrijving'][:50]}")
        else:
            action = add_or_update_row(sheet, item, None)
            added += 1
            print(f"  {action}: {code} - {item['omschrijving'][:50]}")

    # Save Excel
    wb.save(excel_path)
    wb.close()

    print(f"\nDone!")
    print(f"  Added: {added}")
    print(f"  Updated: {updated}")
    print(f"  Total: {len(items)}")


if __name__ == "__main__":
    # Default paths
    tsv_path = "/home/user/vdsboffertes/backend/new_prijzenboek_items.tsv"
    excel_path = "/home/user/vdsboffertes/backend/Juiste opnamelijst.xlsx"

    # Allow command line arguments
    if len(sys.argv) > 1:
        tsv_path = sys.argv[1]
    if len(sys.argv) > 2:
        excel_path = sys.argv[2]

    print("=" * 80)
    print("PRIJZENBOEK TSV IMPORT")
    print("=" * 80)

    import_tsv_to_excel(tsv_path, excel_path, backup=True)

    print("\n" + "=" * 80)

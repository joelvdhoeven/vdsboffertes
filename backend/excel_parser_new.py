"""
Excel prijzenboek parser voor nieuwe structuur
Met ruimte kolommen (C-O) en prijzen in kolommen R-Y
"""
import openpyxl
from typing import List, Dict, Any


def parse_prijzenboek_new(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse de nieuwe prijzenboek structuur
    Ondersteunt twee formaten:
    1. Uitgebreid formaat met ruimte-kolommen (C-O) en prijzen in kolommen R-Y
    2. Simpel formaat: CODERING | OMSCHRIJVING | (lege kolommen) | EENHEID | Materiaal | Uren | Prijs | OMSCHRIJVING OFFERTE

    De parser detecteert automatisch welk formaat wordt gebruikt door de headers te lezen.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet = wb.active

    prijzenboek = []

    # First, read the header row to determine the format
    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = row
        break

    if not header_row:
        wb.close()
        return []

    # Detect format by looking for key header names
    header_map = {}
    for idx, header in enumerate(header_row):
        if header:
            header_str = str(header).strip().upper()
            header_map[header_str] = idx
            # Also check for partial matches
            if "CODERING" in header_str:
                header_map["CODERING"] = idx
            if "EENHEID" in header_str:
                header_map["EENHEID"] = idx
            if "MATERIAAL" in header_str or "MATRIAAL" in header_str:
                header_map["MATERIAAL"] = idx
            if "UREN" in header_str:
                header_map["UREN"] = idx
            if "PRIJS PER STUK" in header_str:
                header_map["PRIJS_PER_STUK"] = idx
            if "OMSCHRIJVING OFFERTE" in header_str:
                header_map["OMSCHRIJVING_OFFERTE"] = idx
            if "OMSCHRIJVING VAKMAN" in header_str:
                header_map["OMSCHRIJVING_VAKMAN"] = idx

    # Determine if this is the simple format (no ruimte columns)
    # Simple format has EENHEID in early columns (before column 10)
    is_simple_format = "EENHEID" in header_map and header_map["EENHEID"] < 10

    row_num = 1  # Will be incremented to 2 at first iteration

    # Use iter_rows for much better performance with read_only mode
    for row in sheet.iter_rows(min_row=2, max_col=25, values_only=True):
        row_num += 1  # Increment at start (2, 3, 4, ...)

        # Get values from tuple - Basis informatie
        code = row[0] if len(row) > 0 else None  # A
        omschrijving = row[1] if len(row) > 1 else None  # B

        # Skip empty rows
        if not code or not omschrijving:
            continue

        if is_simple_format:
            # Simple format: parse based on detected headers
            eenheid_idx = header_map.get("EENHEID", 4)
            materiaal_idx = header_map.get("MATERIAAL", 5)
            uren_idx = header_map.get("UREN", 6)
            prijs_idx = header_map.get("PRIJS_PER_STUK", 7)
            omschrijving_offerte_idx = header_map.get("OMSCHRIJVING_OFFERTE", 8)

            eenheid = row[eenheid_idx] if len(row) > eenheid_idx else None
            materiaal = row[materiaal_idx] if len(row) > materiaal_idx else None
            uren = row[uren_idx] if len(row) > uren_idx else None
            prijs_per_stuk = row[prijs_idx] if len(row) > prijs_idx else None
            omschrijving_offerte = row[omschrijving_offerte_idx] if len(row) > omschrijving_offerte_idx else None

            # Simple format doesn't have ruimte columns or totaal columns
            algemeen_woning = 0.0
            hal_overloop = 0.0
            woonkamer = 0.0
            keuken = 0.0
            toilet = 0.0
            badkamer = 0.0
            slaapk_voor_kl = 0.0
            slaapk_voor_gr = 0.0
            slaapk_achter_kl = 0.0
            slaapk_achter_gr = 0.0
            zolder = 0.0
            berging = 0.0
            meerwerk = 0.0
            totaal = 0.0
            totaal_excl = prijs_per_stuk  # Use prijs_per_stuk as totaal_excl
            totaal_incl = 0.0
        else:
            # Original extended format with ruimte columns
            # Ruimtes (C-O) - indices 2-14
            algemeen_woning = row[2] if len(row) > 2 else None  # C
            hal_overloop = row[3] if len(row) > 3 else None  # D
            woonkamer = row[4] if len(row) > 4 else None  # E
            keuken = row[5] if len(row) > 5 else None  # F
            toilet = row[6] if len(row) > 6 else None  # G
            badkamer = row[7] if len(row) > 7 else None  # H
            slaapk_voor_kl = row[8] if len(row) > 8 else None  # I
            slaapk_voor_gr = row[9] if len(row) > 9 else None  # J
            slaapk_achter_kl = row[10] if len(row) > 10 else None  # K
            slaapk_achter_gr = row[11] if len(row) > 11 else None  # L
            zolder = row[12] if len(row) > 12 else None  # M
            berging = row[13] if len(row) > 13 else None  # N
            meerwerk = row[14] if len(row) > 14 else None  # O

            # Prijzen - indices 16-24 (skip P which is index 15, V which is index 21)
            totaal = row[16] if len(row) > 16 else None  # Q
            eenheid = row[17] if len(row) > 17 else None  # R
            materiaal = row[18] if len(row) > 18 else None  # S
            uren = row[19] if len(row) > 19 else None  # T
            prijs_per_stuk = row[20] if len(row) > 20 else None  # U
            totaal_excl = row[22] if len(row) > 22 else None  # W
            totaal_incl = row[23] if len(row) > 23 else None  # X
            omschrijving_offerte = row[24] if len(row) > 24 else None  # Y

        # Normalize eenheid
        if eenheid:
            eenheid = str(eenheid).strip().lower()

        # Safe float conversion helper
        def safe_float(value):
            if value is None:
                return 0.0
            if isinstance(value, (int, float)):
                return float(value)
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0.0

        item = {
            "code": str(code).strip() if code else "",
            "omschrijving": str(omschrijving).strip(),
            "omschrijving_offerte": str(omschrijving_offerte).strip() if omschrijving_offerte else str(omschrijving).strip(),

            # Ruimtes
            "algemeen_woning": safe_float(algemeen_woning),
            "hal_overloop": safe_float(hal_overloop),
            "woonkamer": safe_float(woonkamer),
            "keuken": safe_float(keuken),
            "toilet": safe_float(toilet),
            "badkamer": safe_float(badkamer),
            "slaapk_voor_kl": safe_float(slaapk_voor_kl),
            "slaapk_voor_gr": safe_float(slaapk_voor_gr),
            "slaapk_achter_kl": safe_float(slaapk_achter_kl),
            "slaapk_achter_gr": safe_float(slaapk_achter_gr),
            "zolder": safe_float(zolder),
            "berging": safe_float(berging),
            "meerwerk": safe_float(meerwerk),

            # Prijzen
            "totaal": safe_float(totaal),
            "eenheid": eenheid or "",
            "materiaal": safe_float(materiaal),
            "uren": safe_float(uren),
            "prijs_per_stuk": safe_float(prijs_per_stuk),
            "totaal_excl": safe_float(totaal_excl),
            "totaal_incl": safe_float(totaal_incl),

            "row_num": row_num
        }

        prijzenboek.append(item)

    wb.close()
    return prijzenboek


if __name__ == "__main__":
    # Test
    prijzenboek = parse_prijzenboek_new("/home/user/vdsboffertes/backend/Juiste opnamelijst.xlsx")
    print(f"Total items: {len(prijzenboek)}")
    print("\nFirst 5 items:")
    for item in prijzenboek[:5]:
        print(f"{item['code']}: {item['omschrijving']} ({item['eenheid']}) - €{item['materiaal']:.2f} + {item['uren']}u")

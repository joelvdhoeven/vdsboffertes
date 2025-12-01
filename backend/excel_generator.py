"""
Excel generator - fills Woonforte offerte template with matched werkzaamheden
Uses Book1.xlsx template and preserves all styling, merged cells, and formatting
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime
from copy import copy


def generate_filled_excel(
    template_path: str,
    matches: List[Dict[str, Any]],
    session_dir: Path
) -> str:
    """
    Generate filled Excel file using Woonforte offerte template
    Preserves template styling and structure

    Args:
        template_path: Path to template (not used, we use our own template)
        matches: List of matched werkzaamheden
        session_dir: Session directory for output

    Returns:
        Path to generated Excel file
    """
    # Use our offerte template
    template_file = Path(__file__).parent / "templates" / "offerte_template.xlsx"

    if not template_file.exists():
        raise FileNotFoundError(f"Template not found: {template_file}")

    # Load template
    wb = load_workbook(template_file)
    ws = wb.active

    # Fill project information in existing cells
    ws['A11'] = "Adres : Project adres"  # TODO: Get from parsed_opname metadata
    ws['H11'] = datetime.now()
    ws['B13'] = "Mireile Onos"  # Default opzichter

    # Delete example data rows (17-20) - we'll replace with real data
    ws.delete_rows(17, 4)

    # Group matches by ruimte
    ruimtes_dict = {}
    for match in matches:
        ruimte = match.get("ruimte", "Algemeen woning")
        if ruimte not in ruimtes_dict:
            ruimtes_dict[ruimte] = []
        ruimtes_dict[ruimte].append(match)

    # Track totals
    totaal_laag_btw = 0.0  # 9% BTW items
    totaal_hoog_btw = 0.0  # 21% BTW items

    # Start inserting at row 17 (after headers at row 16)
    insert_position = 17

    # Define orange fill for ruimte headers (like in template)
    orange_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    # For each ruimte, insert werkzaamheden
    for ruimte_name, ruimte_matches in ruimtes_dict.items():
        # Insert ruimte header row
        ws.insert_rows(insert_position)

        # Set ruimte name and merge cells A:B (like template does)
        ws.cell(row=insert_position, column=1, value=ruimte_name)
        ws.merge_cells(f'A{insert_position}:B{insert_position}')

        # Apply orange styling to ruimte header
        ruimte_cell = ws.cell(row=insert_position, column=1)
        ruimte_cell.fill = orange_fill
        ruimte_cell.font = white_font
        ruimte_cell.alignment = Alignment(horizontal='left', vertical='center')

        insert_position += 1

        # Insert werkzaamheden for this ruimte
        for match in ruimte_matches:
            ws.insert_rows(insert_position)

            code = match["prijzenboek_match"]["code"]
            omschrijving = match["prijzenboek_match"]["omschrijving"]
            hoeveelheid = float(match["opname_item"]["hoeveelheid"])
            eenheid = match["opname_item"]["eenheid"]
            prijs_per_stuk = float(match["prijzenboek_match"].get("prijs_per_stuk", 0))

            # Determine BTW rate (default to 21%)
            btw_rate = 0.21

            # Fill columns according to Book1.xlsx structure:
            # A=Code, B=Omschrijving, C=aantal, D=eenheid
            # E=Prijs/Eenheid 9%, F=Prijs/Eenheid 21%
            # G=Prijs/regel BTW Laag, H=Prijs/regel BTW Hoog, I=Prijs/regel BTW Totaal

            ws.cell(row=insert_position, column=1, value=code)
            ws.cell(row=insert_position, column=2, value=omschrijving)
            ws.cell(row=insert_position, column=3, value=hoeveelheid)
            ws.cell(row=insert_position, column=4, value=eenheid)

            # Calculate regel totaal
            regel_totaal = hoeveelheid * prijs_per_stuk

            if btw_rate == 0.09:
                # 9% BTW (laag tarief)
                ws.cell(row=insert_position, column=5, value=prijs_per_stuk)
                ws.cell(row=insert_position, column=7, value=regel_totaal)
                totaal_laag_btw += regel_totaal
            else:
                # 21% BTW (hoog tarief) - MOST COMMON
                ws.cell(row=insert_position, column=6, value=prijs_per_stuk)
                ws.cell(row=insert_position, column=8, value=regel_totaal)
                totaal_hoog_btw += regel_totaal

            # Total column
            ws.cell(row=insert_position, column=9, value=regel_totaal)

            # Apply number formatting for currency
            for col in [5, 6, 7, 8, 9]:
                ws.cell(row=insert_position, column=col).number_format = '€#,##0.00'

            insert_position += 1

        # Add blank line between ruimtes
        ws.insert_rows(insert_position)
        insert_position += 1

    # Calculate totals
    totaal_excl_btw = totaal_laag_btw + totaal_hoog_btw
    btw_laag = totaal_laag_btw * 0.09
    btw_hoog = totaal_hoog_btw * 0.21
    totaal_incl_btw = totaal_excl_btw + btw_laag + btw_hoog

    # Find and update the Totaal row (should be around current insert_position + 1)
    # Search for "Laag" "Hoog" "Totaal" header row
    totals_header_row = insert_position + 1
    ws.cell(row=totals_header_row, column=7, value="Laag")
    ws.cell(row=totals_header_row, column=8, value="Hoog")
    ws.cell(row=totals_header_row, column=9, value="Totaal")

    # Totaal row - matching Book1.xlsx structure exactly
    totals_row = totals_header_row + 1
    ws.cell(row=totals_row, column=1, value="Totaal")
    ws.cell(row=totals_row, column=2, value=totaal_excl_btw)  # Total in column B (matching Book1.xlsx)
    ws.cell(row=totals_row, column=7, value=totaal_laag_btw)  # Laag BTW total in column G
    ws.cell(row=totals_row, column=8, value=totaal_hoog_btw)  # Hoog BTW total in column H
    ws.cell(row=totals_row, column=9, value=totaal_excl_btw)   # Total in column I

    # Apply formatting
    ws.cell(row=totals_row, column=1).font = Font(bold=True)
    for col in [2, 7, 8, 9]:  # Format columns B, G, H, I
        ws.cell(row=totals_row, column=col).number_format = '€#,##0.00'

    # Totaal incl BTW row
    totals_incl_row = totals_row + 2
    ws.cell(row=totals_incl_row, column=7, value="Totaal incl. BTW")
    ws.cell(row=totals_incl_row, column=9, value=totaal_incl_btw)
    ws.cell(row=totals_incl_row, column=9).number_format = '€#,##0.00'

    # BTW verleggingsregeling section
    btw_section_row = totals_incl_row + 2
    ws.cell(row=btw_section_row, column=1, value="BTW verleggingsregeling")
    ws.cell(row=btw_section_row + 1, column=1, value="BTW verlegd")
    ws.cell(row=btw_section_row + 2, column=1, value="BTW verlegd")

    # Headers
    ws.cell(row=btw_section_row, column=4, value="Grondslag")
    ws.cell(row=btw_section_row, column=5, value="BTW verlegd")

    # 21% BTW
    ws.cell(row=btw_section_row + 1, column=3, value="21%")
    ws.cell(row=btw_section_row + 1, column=4, value=totaal_hoog_btw)
    ws.cell(row=btw_section_row + 1, column=5, value=btw_hoog)

    # 9% BTW
    ws.cell(row=btw_section_row + 2, column=3, value="9%")
    ws.cell(row=btw_section_row + 2, column=4, value=totaal_laag_btw)
    ws.cell(row=btw_section_row + 2, column=5, value=btw_laag)

    # Format BTW section
    for col in [4, 5]:
        ws.cell(row=btw_section_row + 1, column=col).number_format = '€#,##0.00'
        ws.cell(row=btw_section_row + 2, column=col).number_format = '€#,##0.00'

    # Save the filled workbook
    output_filename = f"Offerte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = session_dir / output_filename

    wb.save(str(output_path))
    wb.close()

    print(f"\nGenerated Excel offerte:")
    print(f"  File: {output_path}")
    print(f"  Items: {len(matches)}")
    print(f"  Totaal excl BTW: €{totaal_excl_btw:.2f}")
    print(f"  BTW Laag (9%): €{btw_laag:.2f}")
    print(f"  BTW Hoog (21%): €{btw_hoog:.2f}")
    print(f"  Totaal incl BTW: €{totaal_incl_btw:.2f}")

    return str(output_path)


if __name__ == "__main__":
    # Test the generator
    print("Excel generator ready")
    print(f"Template location: {Path(__file__).parent / 'templates' / 'offerte_template.xlsx'}")
